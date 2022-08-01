import requests
import csv
import time
from openpyxl import Workbook, load_workbook
import os
import csv
import re
import time

#Создание сессии
url = "https://qualysapi.qg2.apps.qualys.eu/api/2.0/fo/session/"

username = input('Введите имя учетной записи с доступами к api:')
password = input('Введите пароль от учетной записи:')

payload='action=login&username=' + str(username)+ '&password=' + str(password)
headers = {
  'X-Requested-With': 'Curl Sample',
  'Content-Type': 'application/x-www-form-urlencoded',
}

s = requests.Session()

response = s.post( url, headers=headers, data=payload)

print(response.text)

#session_id_dict= response.headers['Set-Cookie']

#Обновление списка ip адресов
url_i = "https://qualysapi.qg2.apps.qualys.eu/api/2.0/fo/asset/ip"
payload_i="action=add&enable_vm=1&enable_pc=1&ips=6.6.6.6"

headers = {
  'X-Requested-With': 'Curl',
  'Content-Type': 'application/x-www-form-urlencoded'
}

response_i = s.post( url_i, headers=headers, data=payload_i)
#Выгрузка asset groups

url_1 = "https://qualysapi.qg2.apps.qualys.eu/api/2.0/fo/asset/group/"

payload_1='action=list&output_format=csv&show_attributes=ALL'
headers = {
  'X-Requested-With': 'Curl',
  'Content-Type': 'application/x-www-form-urlencoded'
}

response_1 = s.post( url_1, headers=headers, data=payload_1)


# Сохранение выгрузки в csv файл
file_csv = open(".\\assetgroups_csv.csv", mode='wb')#, encoding='utf-8', newline='')
file_csv.write(response_1.content)
file_csv.close()

# Определение функции работы с csv файлом

def open_csv_file(filename):
    title_l=[]
    ip_l= []
    ID_l = []
    with open(filename, newline='', encoding="utf-8") as File:
        reader = csv.reader(File, delimiter = ",")
        next(File)
        for row in reader:
            if 'BUSINESS_IMPACT' in row:
                for i in range(len(row)):
                    if row[i] == 'TITLE':
                        Title = i
                    if row[i] == 'IP_SET':
                        ip= i
                    if row[i] == 'ID':
                        id= i
                continue
            elif r'END_RESPONSE_BODY' in row[0]:
                break
            print(row[Title])
            print(row[ip])
            title_l.append(row[Title])
            ip_l.append(row[ip])
            ID_l.append(row[id])
    return title_l, ip_l, ID_l

# Превращение рейнджа адресов в массив

def create_ip_list(a,b):
    ip_list = []
    start_i = a.split('.')[3]
    stop_i = b.split('.')[3]
    range_i = int(stop_i)-int(start_i)
    print(range_i)
    ok_1 = a.split('.')[0]
    ok_2 = a.split('.')[1]
    ok_3 = a.split('.')[2]
    for i in range(int(start_i),int(stop_i)+1):
        ip_arr=[]
        ip_arr.append(ok_1)
        ip_arr.append(ok_2)
        ip_arr.append(ok_3)
        ip_arr.append(str(i))
        ip_list.append('.'.join(ip_arr))
    return ip_list

# Проверка директории на наличие файлов с данными

for file in os.listdir():
    if "Server & Network" in file and "~" not in file:
       filename_xl = file
    if "assetgroups_csv" in file:
        filename_csv = file


# Чтение данных из qualys сохранненых в csv файл

asset_title, asset_ip, asset_id = open_csv_file(filename_csv)
asset_ip_list= []
for m in asset_ip:
    asset_ip_list.append(m.split(','))

# выгрузка списка адресов из excel

wb_excel = load_workbook(filename_xl)
ws_excel = wb_excel['Servers']
location=[]
ip_addr=[]
for i in range(450):
    if (ws_excel.cell(row=i+1, column=5).value is not None) and (ws_excel.cell(row=i+1, column=6).value is not None):
        ip_re = re.match('[^\s;,_%^&*]*', ws_excel.cell(row=i+1, column=5).value)
        pars_addr = ip_re.group(0)
        location.append(ws_excel.cell(row=i+1, column=6).value)
        ip_addr.append(pars_addr)



wb_excel.save(filename_xl)


#Парсинг адресов из qualys

new_asset_ip_list = []
for km in asset_ip_list:
    test = km
    km =[]
    for m in range(len(test)):
        if '-' in test[m]:
            var = test[m]
            expand_range = create_ip_list(*var.split('-'))
            #range_ip = var.split('-')
            km = km + expand_range
        else:
            km.append(test[m])
    new_asset_ip_list.append(km)
    

# Сравнение адресов из qualys с файлом из excel (добавление отсутствующих)
for m in range(len(asset_title)):
    for i in range(len(location)):
        if location[i] in asset_title[m]:
            if ip_addr[i] not in new_asset_ip_list[m]:
                new_asset_ip_list[m].append(ip_addr[i])


# Сравнение адресов из qualys с файлом из excel, удаление старых

for m in range(len(asset_title)):
    for k in new_asset_ip_list[m]:
        if k not in ip_addr:
            new_asset_ip_list[m].remove(k)
            
#Наполнение хостов
            '''
ip_addr.remove('IP')
ip_addr_to_host = ','.join(ip_addr)
            
url_i = "https://qualysapi.qg2.apps.qualys.eu/api/2.0/fo/asset/ip/?action=update"
payload_i='enable_vm=1&ips=' + str(ip_addr_to_host)

headers = {
  'X-Requested-With': 'Curl',
  'Content-Type': 'application/x-www-form-urlencoded'
}
print(payload_i)
response_i = s.post( url_i, headers=headers, data=payload_i)            
        '''
# Связь с API отправка нового скоупа адресов в qualys

for m in range(len(new_asset_ip_list)):
    if 'All' not in asset_title[m]:
        if len(new_asset_ip_list[m])== 0:
            asset_ip_parameter= ','.join(new_asset_ip_list[m])
        elif new_asset_ip_list[m][0] is '':
            asset_ip_parameter= ','.join(new_asset_ip_list[m][1:])
        else:
            asset_ip_parameter= ','.join(new_asset_ip_list[m])
        url_i = "https://qualysapi.qg2.apps.qualys.eu/api/2.0/fo/asset/ip/?action=add"
        payload_i='enable_vm=1&ips=' + str(asset_ip_parameter)
        response_i = s.post( url_i, headers=headers, data=payload_i)
        print(response_i.text)
        time.sleep(3)
    else:
        pass
        
for m in range(len(new_asset_ip_list)):
    if 'wave' not in asset_title[m]:
        if len(new_asset_ip_list[m])== 0:
            asset_ip_parameter= ','.join(new_asset_ip_list[m])
        elif new_asset_ip_list[m][0] is '':
            asset_ip_parameter= ','.join(new_asset_ip_list[m][1:])
        else:
            asset_ip_parameter= ','.join(new_asset_ip_list[m])
        url_2 = "https://qualysapi.qg2.apps.qualys.eu/api/2.0/fo/asset/group/?action=edit"    
        payload_2='id=' + str(asset_id[m]) + '&set_ips='+str(asset_ip_parameter)
        headers = {
          'X-Requested-With': 'Curl',
          'Content-Type': 'application/x-www-form-urlencoded'
        }
        #print(payload_2)
        time.sleep(3)
        
        response_2 = s.post( url_2, headers=headers, data=payload_2)
        time.sleep(1)
        if 'Asset Group Updated Successfully' in response_2.text:
            print(str(asset_title[m]) + ' has updated succesfully!')
        else:
            print(str(asset_title[m]) + ' has an issue while updating!')
            print(str(response_2.text))
    else:
        pass





        

    #print(payload_2)
    #response_2 = s.post( url_2, headers=headers, data=payload_2)

    
'''    
    if 'Asset Group Updated Successfully' in response_2.text:
        print(str(asset_title[m]) + ' has updated succesfully!')
    else:
        print(str(asset_title[m]) + ' has an issue while updating!')
    #'Asset Group Updated Successfully' in response_2.text
'''    
